import openpyxl
import sys

def leer_excel(ruta_excel):
    try:
        workbook = openpyxl.load_workbook(ruta_excel)
        sheet = workbook.active
        print(f"El Excel tiene {sheet.max_row} filas y {sheet.max_column} columnas")
        print("=" * 50)
        
        # Mostrar encabezados
        print("\n--- ENCABEZADOS ---")
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            print(f"Columna {col}: {header}")
        
        # Mostrar primeras filas de datos
        print("\n--- PRIMERAS 5 FILAS DE DATOS ---")
        for row in range(2, min(7, sheet.max_row + 1)):
            print(f"\nFila {row}:")
            for col in range(1, sheet.max_column + 1):
                value = sheet.cell(row=row, column=col).value
                header = sheet.cell(row=1, column=col).value
                print(f"  {header}: {value}")
                
    except FileNotFoundError:
        print(f"Error: No se encuentra el archivo en la ruta: {ruta_excel}")
    except Exception as e:
        print(f"Error al leer el Excel: {str(e)}")

if __name__ == "__main__":
    ruta = r"C:\Mepryl4.2\EXPORTACION DICTAMENES AL 30-06-2026.xlsx"
    leer_excel(ruta)
